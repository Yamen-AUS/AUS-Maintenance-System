"""
AUS Maintenance System - Professional Excel Generator
Creates a complete Excel workbook with multiple sheets, charts, formulas, and formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import json

# تحميل البيانات
with open('AUS_Maintenance_Data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
    tasks = data['tasks']

# إنشاء Workbook
wb = Workbook()
wb.remove(wb.active)  # حذف الورقة الافتراضية

# ========== الألوان والأنماط ==========
COLORS = {
    'primary': '4472C4',      # أزرق
    'success': '70AD47',      # أخضر
    'warning': 'FFC000',      # أصفر
    'danger': 'E74C3C',       # أحمر
    'info': '5DADE2',         # أزرق فاتح
    'dark': '2C3E50',         # رمادي غامق
    'light': 'ECF0F1',        # رمادي فاتح
    'white': 'FFFFFF'
}

# تنسيقات مشتركة
header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
title_font = Font(name='Calibri', size=18, bold=True, color=COLORS['primary'])
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ========== Sheet 1: DASHBOARD ==========
ws_dash = wb.create_sheet('📊 Dashboard', 0)
ws_dash.sheet_properties.tabColor = COLORS['primary']

# العنوان الرئيسي
ws_dash['B2'] = 'AUS MAINTENANCE SYSTEM'
ws_dash['B2'].font = Font(name='Calibri', size=24, bold=True, color=COLORS['primary'])
ws_dash['B3'] = 'Arab Unity School - Academic Year 2025-2026'
ws_dash['B3'].font = Font(name='Calibri', size=12, color=COLORS['dark'])

# KPIs
completed = len([t for t in tasks if 'Completed' in t['status']])
pending = len(tasks) - completed
total_cost = sum(t['cost'] for t in tasks)
avg_cost = total_cost / len(tasks)

kpi_data = [
    ['METRIC', 'VALUE', 'STATUS'],
    ['Total Tasks', len(tasks), '📊'],
    ['Completed', completed, '✅'],
    ['Pending', pending, '⏳'],
    ['Completion Rate', f"{(completed/len(tasks)*100):.1f}%", '📈'],
    ['Total Cost (AED)', f"{total_cost:,.0f}", '💰'],
    ['Average Cost', f"{avg_cost:,.0f}", '📊']
]

# كتابة KPIs
start_row = 6
for i, row in enumerate(kpi_data):
    for j, value in enumerate(row):
        cell = ws_dash.cell(row=start_row+i, column=2+j, value=value)
        if i == 0:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if j == 1:
                cell.font = Font(size=14, bold=True)

# تنسيق الأعمدة
ws_dash.column_dimensions['B'].width = 20
ws_dash.column_dimensions['C'].width = 15
ws_dash.column_dimensions['D'].width = 10

# ========== Sheet 2: TASKS ==========
ws_tasks = wb.create_sheet('📋 Tasks Data')
ws_tasks.sheet_properties.tabColor = COLORS['success']

# العناوين
headers = ['ID', 'Title', 'Category', 'Location', 'Term', 'Priority', 'Status', 
           'Assigned To', 'Reported By', 'Date Reported', 'Due Date', 'Completed', 
           'Cost', 'Vendor', 'Progress', 'Notes']

for col, header in enumerate(headers, 1):
    cell = ws_tasks.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# البيانات
for row_idx, task in enumerate(tasks, 2):
    ws_tasks.cell(row=row_idx, column=1, value=task['id'])
    ws_tasks.cell(row=row_idx, column=2, value=task['title'])
    ws_tasks.cell(row=row_idx, column=3, value=task['category'])
    ws_tasks.cell(row=row_idx, column=4, value=task['location'])
    ws_tasks.cell(row=row_idx, column=5, value=task['term'])
    ws_tasks.cell(row=row_idx, column=6, value=task['priority'].replace('🔴', '').replace('🟠', '').replace('🟡', '').strip())
    ws_tasks.cell(row=row_idx, column=7, value=task['status'].replace('✅', '').replace('⏳', '').strip())
    ws_tasks.cell(row=row_idx, column=8, value=task['assignedTo'])
    ws_tasks.cell(row=row_idx, column=9, value=task['reportedBy'])
    ws_tasks.cell(row=row_idx, column=10, value=task['dateReported'])
    ws_tasks.cell(row=row_idx, column=11, value=task['dueDate'])
    ws_tasks.cell(row=row_idx, column=12, value=task.get('dateCompleted', ''))
    ws_tasks.cell(row=row_idx, column=13, value=task['cost']).number_format = '#,##0'
    ws_tasks.cell(row=row_idx, column=14, value=task['vendor'])
    ws_tasks.cell(row=row_idx, column=15, value=task['progress'])
    ws_tasks.cell(row=row_idx, column=16, value=task['notes'])
    
    # تنسيق الصفوف
    for col in range(1, 17):
        ws_tasks.cell(row=row_idx, column=col).border = border
        ws_tasks.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')

# تنسيق الأعمدة
column_widths = [12, 25, 18, 18, 10, 12, 12, 18, 18, 14, 14, 14, 12, 20, 10, 45]
for i, width in enumerate(column_widths, 1):
    ws_tasks.column_dimensions[get_column_letter(i)].width = width

# ========== Sheet 3: ANALYTICS ==========
ws_analytics = wb.create_sheet('📈 Analytics')
ws_analytics.sheet_properties.tabColor = COLORS['info']

# تحليل الحالات
ws_analytics['B2'] = 'STATUS ANALYSIS'
ws_analytics['B2'].font = title_font

status_data = [
    ['Status', 'Count', 'Percentage'],
    ['Completed', completed, f"{(completed/len(tasks)*100):.1f}%"],
    ['Pending', pending, f"{(pending/len(tasks)*100):.1f}%"]
]

for i, row in enumerate(status_data, 4):
    for j, value in enumerate(row, 2):
        cell = ws_analytics.cell(row=i, column=j, value=value)
        if i == 4:
            cell.font = header_font
            cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

# Pie Chart للحالات
pie = PieChart()
pie.title = "Task Status Distribution"
pie.height = 10
pie.width = 15

labels = Reference(ws_analytics, min_col=2, min_row=5, max_row=6)
data = Reference(ws_analytics, min_col=3, min_row=4, max_row=6)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)

ws_analytics.add_chart(pie, "F4")

# تحليل التكلفة حسب الفئة
ws_analytics['B12'] = 'COST BY CATEGORY'
ws_analytics['B12'].font = title_font

categories = {}
for task in tasks:
    cat = task['category']
    if cat not in categories:
        categories[cat] = 0
    categories[cat] += task['cost']

cat_data = [['Category', 'Total Cost (AED)']]
for cat, cost in sorted(categories.items(), key=lambda x: x[1], reverse=True):
    cat_data.append([cat, cost])

for i, row in enumerate(cat_data, 14):
    for j, value in enumerate(row, 2):
        cell = ws_analytics.cell(row=i, column=j, value=value)
        if i == 14:
            cell.font = header_font
            cell.fill = header_fill
        else:
            if j == 3:
                cell.number_format = '#,##0'
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

# Bar Chart للتكلفة
bar = BarChart()
bar.type = "col"
bar.title = "Cost by Category"
bar.height = 12
bar.width = 20

cats = Reference(ws_analytics, min_col=2, min_row=15, max_row=14+len(categories))
values = Reference(ws_analytics, min_col=3, min_row=14, max_row=14+len(categories))
bar.add_data(values, titles_from_data=True)
bar.set_categories(cats)

ws_analytics.add_chart(bar, "F14")

# ========== Sheet 4: SUMMARY ==========
ws_summary = wb.create_sheet('📑 Summary')
ws_summary.sheet_properties.tabColor = COLORS['warning']

summary_data = [
    ['AUS MAINTENANCE SYSTEM - EXECUTIVE SUMMARY'],
    [''],
    ['Report Date:', '2026-03-14'],
    ['Academic Year:', '2025-2026'],
    [''],
    ['OVERVIEW'],
    ['Total Number of Tasks:', len(tasks)],
    ['Completed Tasks:', completed],
    ['Pending Tasks:', pending],
    ['Completion Rate:', f"{(completed/len(tasks)*100):.1f}%"],
    [''],
    ['FINANCIAL SUMMARY'],
    ['Total Project Cost:', f"AED {total_cost:,.0f}"],
    ['Completed Cost:', f"AED {sum(t['cost'] for t in tasks if 'Completed' in t['status']):,.0f}"],
    ['Pending Cost:', f"AED {sum(t['cost'] for t in tasks if 'Completed' not in t['status']):,.0f}"],
    ['Average Cost per Task:', f"AED {avg_cost:,.0f}"],
    [''],
    ['PRIORITY BREAKDOWN'],
    ['Critical:', len([t for t in tasks if 'Critical' in t['priority']])],
    ['High:', len([t for t in tasks if 'High' in t['priority']])],
    ['Medium:', len([t for t in tasks if 'Medium' in t['priority']])],
    ['Low:', len([t for t in tasks if 'Low' in t['priority']])]
]

for i, row in enumerate(summary_data, 2):
    if isinstance(row, list) and len(row) > 0:
        ws_summary.cell(row=i, column=2, value=row[0])
        if len(row) > 1:
            ws_summary.cell(row=i, column=3, value=row[1])
        
        # تنسيق العناوين
        if row[0] in ['AUS MAINTENANCE SYSTEM - EXECUTIVE SUMMARY', 'OVERVIEW', 'FINANCIAL SUMMARY', 'PRIORITY BREAKDOWN']:
            ws_summary.cell(row=i, column=2).font = title_font
            ws_summary.merge_cells(f'B{i}:D{i}')

ws_summary.column_dimensions['B'].width = 30
ws_summary.column_dimensions['C'].width = 20

# حفظ الملف
wb.save('AUS_Maintenance_Professional.xlsx')
print('✅ تم إنشاء الملف: AUS_Maintenance_Professional.xlsx')
